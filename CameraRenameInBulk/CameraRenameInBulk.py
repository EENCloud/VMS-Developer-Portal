import requests
import argparse
import os
from openpyxl import load_workbook
from tqdm import tqdm

# Create a session
s = requests.Session()

def read_camera_data(excel_file_path):
    camera_data = []

    workbook = load_workbook(filename=excel_file_path, read_only=True)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2:
            camera_id = str(row[0]).strip()
            new_name = str(row[1]).strip()
            camera_data.append((camera_id, new_name))
            print(f"Read camera_id: {camera_id}, new_name: {new_name}")
        else:
            print(f"Skipped row due to insufficient values: {row}")

    workbook.close()
    return camera_data

def update_camera_names_and_write_result(camera_data, baseurl, access_token, token_type, excel_file_path):
    workbook = load_workbook(filename=excel_file_path)
    sheet = workbook.active

    # Assuming the status will be written to the third column
    sheet.cell(row=1, column=3, value="Status")

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "authorization": f"{token_type} {access_token}"
    }

    for index, (camera_id, new_name) in enumerate(tqdm(camera_data, desc="Updating Camera Names", ncols=100), start=2):  # start=2 to skip the header row
        data = {
            'id': camera_id,
            'name': new_name
        }
        url = f"https://{baseurl}/api/v3.0/cameras/{camera_id}"
        response = s.patch(url, headers=headers, json=data)

        if response.status_code == 204:
            print(f"Camera {camera_id} name updated to {new_name}")
            sheet.cell(row=index, column=3, value="Successfully renamed")
        else:
            print(f"Failed to update camera {camera_id} name")
            sheet.cell(row=index, column=3, value=f"Error renaming. HTTP {response.status_code}: {response.text}")

    # Save the workbook with the results
    workbook.save(excel_file_path)
    workbook.close()

    print("Camera name update process complete.")

def main():
    parser = argparse.ArgumentParser(description="Update camera names using the given Excel file and access token.")
    parser.add_argument('-f', '--file', required=True, help="Path to Excel file containing camera data")
    parser.add_argument('-t', '--token', required=True, help="Access token for authorization")

    args = parser.parse_args()

    if not os.path.isfile(args.file):
        print("The provided Excel file path does not exist. Please provide a valid path.")
        return

    excel_file_path = args.file

    # Read camera data from Excel file
    camera_data = read_camera_data(excel_file_path)

    # Access token
    access_token = args.token
    token_type = "Bearer"

    # Get client settings
    url = "https://api.eagleeyenetworks.com/api/v3.0/clientSettings"
    headers = {
        "accept": "application/json",
        "authorization": f"{token_type} {access_token}"
    }
    response = s.get(url, headers=headers)
    baseurl = response.json()["httpsBaseUrl"]["hostname"]

    # Update camera names and write results
    update_camera_names_and_write_result(camera_data, baseurl, access_token, token_type, excel_file_path)

if __name__ == "__main__":
    main()
